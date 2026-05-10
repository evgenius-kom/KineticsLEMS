"""Cross-validation against the ABPOT/DAPOT reference α(T) tables in
``theory/Вязовкин.xlsx``.

Each material sheet has several heating rates side-by-side as (α, T) column
pairs, plus a column with reference E_a(α) values from an independent sheet.
We rebuild ``ConversionRun`` objects directly from those tables (skipping
our DSC pre-processing, which expects raw y(T)) and compare our methods'
output against the reference column.

Run from repo root: ``.venv/bin/python scripts/validate_abpot_dapot.py``.
"""
from __future__ import annotations

from pathlib import Path

import numpy as np
from openpyxl import load_workbook

from kinetics_lems.constants import SEC_PER_MIN
from kinetics_lems.conversion import ConversionRun
from kinetics_lems.methods import friedman, kas, kissinger, ofw, vyazovkin, vyazovkin_aic

REPO = Path(__file__).resolve().parents[1]


def _read_runs(sheet, rate_columns: dict[float, tuple[int, int]]) -> list[ConversionRun]:
    """Load (α, T) pairs from given column pairs and build ConversionRun objects.

    rate_columns: {β (K/min) -> (col_alpha, col_T), 1-based}.
    """
    runs: list[ConversionRun] = []
    for beta, (c_alpha, c_T) in rate_columns.items():
        a_list: list[float] = []
        T_list: list[float] = []
        for r in range(3, sheet.max_row + 1):  # data starts at row 3
            a = sheet.cell(r, c_alpha).value
            T = sheet.cell(r, c_T).value
            if a is None or T is None:
                continue
            try:
                a_f = float(a)
                T_f = float(T)
            except (TypeError, ValueError):
                continue
            a_list.append(a_f)
            T_list.append(T_f)
        if not a_list:
            continue
        alpha = np.asarray(a_list)
        T = np.asarray(T_list)
        # Sort by α and de-duplicate just in case.
        order = np.argsort(alpha)
        alpha = alpha[order]
        T = T[order]
        # Clip α into [0, 1] (xlsx had 1.00013 in one place).
        alpha = np.clip(alpha, 0.0, 1.0)

        # dα/dt = (dα/dT) · β_per_sec.
        beta_per_sec = beta / SEC_PER_MIN
        dadT = np.gradient(alpha, T)
        runs.append(
            ConversionRun(
                rate_K_per_min=float(beta),
                temperature=T,
                alpha=alpha,
                dalpha_dt=dadT * beta_per_sec,
            )
        )
    return runs


def _read_reference_Ea(sheet, col: int) -> tuple[np.ndarray, np.ndarray]:
    """Read the (α, E_a) reference column as a sparse table indexed by data rows."""
    alphas: list[float] = []
    Eas: list[float] = []
    for r in range(3, sheet.max_row + 1):
        # Reference α is the same column 1 in both sheets.
        a = sheet.cell(r, 1).value
        e = sheet.cell(r, col).value
        if a is None or e is None:
            continue
        try:
            a_f = float(a)
            e_f = float(e)
        except (TypeError, ValueError):
            continue
        if not np.isfinite(e_f):
            continue
        alphas.append(a_f)
        Eas.append(e_f)
    return np.asarray(alphas), np.asarray(Eas)


def _print_table(label: str, alphas: np.ndarray, ref_Ea: np.ndarray, results: dict):
    print(f"\n=== {label} ===")
    header = f"{'α':>6} | {'ref Ea':>8} | " + " | ".join(f"{n:>10}" for n in results)
    print(header)
    print("-" * len(header))
    for i, a in enumerate(alphas):
        row = [f"{a:6.2f}", f"{ref_Ea[i]:8.2f}"]
        for ea in results.values():
            row.append(f"{ea[i]:10.2f}" if np.isfinite(ea[i]) else f"{'nan':>10}")
        print(" | ".join(row))


def _evaluate(label: str, runs: list[ConversionRun], ref_alphas: np.ndarray,
              ref_Ea_kJ: np.ndarray):
    """Run all methods on the given runs and print comparison vs reference."""
    # Restrict to alphas where Vyazovkin AIC's window is valid (≥ Δα from edges).
    delta = 0.02
    mask = (ref_alphas >= delta) & (ref_alphas <= 1.0 - delta)
    alphas = ref_alphas[mask]
    ref_Ea = ref_Ea_kJ[mask]

    fri = friedman(runs, alphas)
    kas_ = kas(runs, alphas)
    ofw_ = ofw(runs, alphas)
    vya = vyazovkin(runs, alphas)
    aic = vyazovkin_aic(runs, alphas, delta_alpha=delta)
    kiss = kissinger(runs)

    results = {
        "Friedman": fri.Ea_kJ_per_mol,
        "KAS": kas_.Ea_kJ_per_mol,
        "OFW": ofw_.Ea_kJ_per_mol,
        "Vyazov": vya.Ea_kJ_per_mol,
        "Vya-AIC": aic.Ea_kJ_per_mol,
    }
    _print_table(label, alphas, ref_Ea, results)

    # Summary metrics: mean abs diff and max abs diff for each method.
    print(f"\n  rates: {[r.rate_K_per_min for r in runs]} K/min")
    print(f"  α points compared: {alphas.size}")
    for n, ea in results.items():
        valid = np.isfinite(ea)
        diff = ea[valid] - ref_Ea[valid]
        print(
            f"  {n:8s}  mean abs Δ = {np.mean(np.abs(diff)):6.2f} kJ/mol  "
            f"|  max abs Δ = {np.max(np.abs(diff)):6.2f}  |  mean Ea = {np.mean(ea[valid]):6.2f}"
        )
    print(f"  Kissinger    Ea = {kiss.Ea_kJ_per_mol:6.2f} kJ/mol  (R²={kiss.r_squared:.4f})  "
          f"(reference is per-α, not directly comparable)")


def main():
    wb = load_workbook(REPO / "theory" / "Вязовкин.xlsx", data_only=True)

    # ---- ABPOT ----
    abpot = wb["АБПОТ"]
    # Row 1 labels: 0.2, 1, 2.5, 5 (K/min). Each rate = (col_alpha, col_T)=(c, c+1).
    abpot_rates = {
        0.2: (1, 2),
        1.0: (4, 5),
        2.5: (7, 8),
        5.0: (10, 11),
    }
    abpot_runs = _read_runs(abpot, abpot_rates)
    abpot_ref_a, abpot_ref_Ea = _read_reference_Ea(abpot, col=13)
    _evaluate("ABPOT", abpot_runs, abpot_ref_a, abpot_ref_Ea)

    # ---- DAPOT ----
    dapot = wb["ДАПОТ"]
    # Row 1 labels: <unlabeled>, 2.5, 5, 10. The unlabeled first column pair
    # — based on T(α=0.01) being lower than at 2.5 K/min — is almost
    # certainly 1 K/min. Use only the three labeled rates to be safe.
    dapot_rates = {
        2.5: (4, 5),
        5.0: (7, 8),
        10.0: (10, 11),
    }
    dapot_runs = _read_runs(dapot, dapot_rates)
    dapot_ref_a, dapot_ref_Ea = _read_reference_Ea(dapot, col=13)
    _evaluate("DAPOT (3 labeled rates: 2.5, 5, 10 K/min)", dapot_runs, dapot_ref_a, dapot_ref_Ea)


if __name__ == "__main__":
    main()
