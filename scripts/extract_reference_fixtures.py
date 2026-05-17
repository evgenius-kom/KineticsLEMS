"""One-shot extraction of all reference data from ``theory/extra/*.xlsx``
into ``data/reference_workbooks/`` so the xlsx files can be deleted.

Produces the following layout::

    data/reference_workbooks/
        abpot/                         # ABPOT polymer, 4 heating rates
            rate_0.2.tsv  rate_1.0.tsv  rate_2.5.tsv  rate_5.0.tsv
            reference_Ea.tsv  README.md
        dapot/                         # DAPOT polymer, 4 heating rates
            rate_1.0.tsv  rate_2.5.tsv  rate_5.0.tsv  rate_10.0.tsv
            reference_Ea.tsv  README.md
        epoxy_pv15/                    # Epoxy PV15-0,5, 3 heating rates + ref E_a
            rate_2.5.tsv  rate_5.0.tsv  rate_10.0.tsv
            reference_Ea.tsv  README.md
        skm_prepreg/                   # SKM preimpregnated polymer, 6 heating rates
            rate_1.0.tsv  rate_2.5.tsv  rate_5.0.tsv
            rate_10.0.tsv rate_20.0.tsv rate_40.0.tsv
            README.md
        xlsx_crosscheck/               # numbers precomputed by xlsx formulas
            zplot_abpot_5kpm.tsv
            zplot_dapot_5kpm.tsv
            A_calculated_abpot.tsv
            reaction_order_abpot.tsv
            reaction_order_dapot.tsv
            README.md

Run from repo root::

    .venv/bin/python scripts/extract_reference_fixtures.py
"""
from __future__ import annotations

from pathlib import Path
from typing import NamedTuple

import numpy as np
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

REPO = Path(__file__).resolve().parents[1]
SRC = REPO / "theory" / "extra"
OUT = REPO / "data" / "reference_workbooks"

VYAZOVKIN_XLSX = SRC / "Вязовкин.xlsx"
RASCHET_XLSM = SRC / "для расчета по Вязовкину.xlsm"
SVODKA_XLSX = SRC / "Сводная таблица.xlsx"


# ---------- Generic helpers ----------

def _read_pairs(
    sheet: Worksheet, c_alpha: int, c_T: int, start_row: int = 3
) -> tuple[np.ndarray, np.ndarray]:
    """Read (α, T) columns. Skip rows where either cell is missing/non-numeric."""
    alpha: list[float] = []
    T: list[float] = []
    for r in range(start_row, sheet.max_row + 1):
        a = sheet.cell(r, c_alpha).value
        t = sheet.cell(r, c_T).value
        if a is None or t is None:
            continue
        try:
            a_f, t_f = float(a), float(t)
        except (TypeError, ValueError):
            continue
        alpha.append(a_f)
        T.append(t_f)
    a_arr = np.asarray(alpha)
    t_arr = np.asarray(T)
    order = np.argsort(a_arr)
    a_arr = np.clip(a_arr[order], 0.0, 1.0)
    t_arr = t_arr[order]
    return a_arr, t_arr


def _read_reference_Ea(
    sheet: Worksheet, alpha_col: int, ea_col: int, start_row: int = 3
) -> tuple[np.ndarray, np.ndarray]:
    """Read per-α reference E_a from a worksheet column."""
    alpha: list[float] = []
    Ea: list[float] = []
    for r in range(start_row, sheet.max_row + 1):
        a = sheet.cell(r, alpha_col).value
        e = sheet.cell(r, ea_col).value
        if a is None or e is None:
            continue
        try:
            a_f, e_f = float(a), float(e)
        except (TypeError, ValueError):
            continue
        if not np.isfinite(e_f):
            continue
        alpha.append(a_f)
        Ea.append(e_f)
    return np.asarray(alpha), np.asarray(Ea)


def _save_tsv(path: Path, headers: list[str], columns: list[np.ndarray]) -> None:
    """Write columns as a TSV using %.10g for floats (full round-trip precision,
    compact for both normal-range and very-small/very-large values)."""
    path.parent.mkdir(parents=True, exist_ok=True)
    n = len(columns[0])
    for c in columns:
        assert len(c) == n, "column length mismatch"
    with path.open("w", encoding="utf-8") as f:
        f.write("\t".join(headers) + "\n")
        for i in range(n):
            row = [
                f"{float(c[i]):.10g}" if isinstance(c[i], (int, float, np.floating, np.integer)) else str(c[i])
                for c in columns
            ]
            f.write("\t".join(row) + "\n")


def _write_readme(path: Path, content: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(content.strip() + "\n", encoding="utf-8")


# ---------- Per-material extractors ----------

class MaterialSpec(NamedTuple):
    folder: str
    sheet: str
    rates: dict[float, tuple[int, int]]  # rate K/min -> (alpha_col, T_col)
    ea_cols: tuple[int, int] | None       # (alpha_col, ea_col) or None
    readme: str


def extract_material(xlsx: Path, spec: MaterialSpec) -> None:
    wb = load_workbook(xlsx, data_only=True)
    sheet = wb[spec.sheet]
    out = OUT / spec.folder
    for beta, (c_a, c_t) in spec.rates.items():
        a, T = _read_pairs(sheet, c_a, c_t)
        _save_tsv(out / f"rate_{beta}.tsv", ["alpha", "T_K"], [a, T])
    if spec.ea_cols is not None:
        a_ref, ea_ref = _read_reference_Ea(sheet, *spec.ea_cols)
        _save_tsv(out / "reference_Ea.tsv", ["alpha", "Ea_kJ_per_mol"], [a_ref, ea_ref])
    _write_readme(out / "README.md", spec.readme)
    wb.close()


_ABPOT = MaterialSpec(
    folder="abpot",
    sheet="АБПОТ",
    rates={0.2: (1, 2), 1.0: (4, 5), 2.5: (7, 8), 5.0: (10, 11)},
    ea_cols=(1, 13),
    readme="""
# ABPOT polymer

Source: `theory/extra/Вязовкин.xlsx` sheet `АБПОТ`.

## Data files

| File | Columns | Notes |
|------|---------|-------|
| `rate_0.2.tsv`, `rate_1.0.tsv`, `rate_2.5.tsv`, `rate_5.0.tsv` | `alpha`, `T_K` | (α, T) pairs per heating rate, sorted by α, clipped to [0, 1] |
| `reference_Ea.tsv` | `alpha`, `Ea_kJ_per_mol` | per-α reference E_a from an independent Excel-based Vyazovkin computation |

Heating rates: 0.2, 1.0, 2.5, 5.0 K/min. α grid step ≈ 0.01.

## Applicable tests

* **Integral-method validation** — KAS / OFW / Vyazovkin reproduce
  the per-α reference E_a (see [`tests/test_reference_abpot_dapot.py`](../../../tests/test_reference_abpot_dapot.py)).
* **Multi-rate robustness** — 4-rate dataset; useful for sanity-checking
  Vyazovkin and Vyazovkin-AIC.
* **Z(α) cross-check at 5 K/min** — paired with
  [`../xlsx_crosscheck/zplot_abpot_5kpm.tsv`](../xlsx_crosscheck/).
* **A(α) cross-check at rates 1, 2.5, 5** — paired with
  [`../xlsx_crosscheck/A_calculated_abpot.tsv`](../xlsx_crosscheck/).
* **Reaction-order cross-check at 5 K/min** — paired with
  [`../xlsx_crosscheck/reaction_order_abpot.tsv`](../xlsx_crosscheck/).
""",
)


_DAPOT = MaterialSpec(
    folder="dapot",
    sheet="ДАПОТ",
    # The Вязовкин.xlsx ДАПОТ sheet has rates 2.5/5/10. The 1 K/min run was
    # extracted separately from "для расчета по Вязовкину.xlsm" (3 скорости
    # ДАПОТ sheet); see extract_dapot_rate_1() below for that path.
    rates={2.5: (4, 5), 5.0: (7, 8), 10.0: (10, 11)},
    ea_cols=(1, 13),
    readme="""
# DAPOT polymer

Source: `theory/extra/Вязовкин.xlsx` sheet `ДАПОТ`
(plus β = 1 K/min from `для расчета по Вязовкину.xlsm` sheet `3 скорости ДАПОТ`).

## Data files

| File | Columns | Notes |
|------|---------|-------|
| `rate_1.0.tsv`, `rate_2.5.tsv`, `rate_5.0.tsv`, `rate_10.0.tsv` | `alpha`, `T_K` | (α, T) pairs per heating rate |
| `reference_Ea.tsv` | `alpha`, `Ea_kJ_per_mol` | per-α reference E_a from independent Excel Vyazovkin |

Heating rates: 1.0, 2.5, 5.0, 10.0 K/min. α grid step ≈ 0.01.

## Applicable tests

* **Integral-method validation** — KAS / OFW / Vyazovkin against reference E_a
  ([`tests/test_reference_abpot_dapot.py`](../../../tests/test_reference_abpot_dapot.py)).
* **Z(α) cross-check at 5 K/min** ([`../xlsx_crosscheck/zplot_dapot_5kpm.tsv`](../xlsx_crosscheck/)).
* **Reaction-order cross-check at 5 K/min** ([`../xlsx_crosscheck/reaction_order_dapot.tsv`](../xlsx_crosscheck/)).
""",
)


_EPOXY = MaterialSpec(
    folder="epoxy_pv15",
    sheet="Эпоксидка PV15-0,5",
    rates={2.5: (1, 2), 5.0: (3, 3), 10.0: (4, 4)},  # corrected below
    ea_cols=(1, 6),
    readme="""
# Epoxy PV15-0,5

Source: `theory/extra/Вязовкин.xlsx` sheet `Эпоксидка PV15-0,5`.

Epoxy resin with low (E_a ≈ 50 kJ/mol) activation energy that
*decreases* with α — classic Avrami-style autocatalytic-looking
fingerprint that is interesting because it is *different* from the
ABPOT/DAPOT polymers which have flat E_a(α).

## Data files

| File | Columns | Notes |
|------|---------|-------|
| `rate_2.5.tsv`, `rate_5.0.tsv`, `rate_10.0.tsv` | `alpha`, `T_K` | (α, T) per heating rate |
| `reference_Ea.tsv` | `alpha`, `Ea_kJ_per_mol` | independent Excel Vyazovkin reference |

Heating rates: 2.5, 5.0, 10.0 K/min. α grid step ≈ 0.01.

## Applicable tests

* **Integral-method validation** under declining E_a(α) — KAS / OFW /
  Vyazovkin should track the reference even though the E curve is not
  flat. Catches drift on materials where multi-step character is mild.
* **Reaction-model identification** — Z(α) master-plot should rank
  Avrami / nth-order higher than diffusion models, which the reference
  E_a profile expects.
""",
)


_SKM = MaterialSpec(
    folder="skm_prepreg",
    sheet="прпепрег-СКМ",
    # Sheet layout: column 1 = α; columns 2–7 = T per rate.
    rates={1.0: (1, 2), 2.5: (1, 3), 5.0: (1, 4), 10.0: (1, 5), 20.0: (1, 6), 40.0: (1, 7)},
    ea_cols=None,  # no reference E_a in this sheet
    readme="""
# SKM preimpregnated polymer

Source: `theory/extra/для расчета по Вязовкину.xlsm` sheet `прпепрег-СКМ`.

Six heating rates: 1, 2.5, 5, 10, 20, 40 K/min — span of 40× between
slowest and fastest. The widest β-range in this corpus, useful for
checking multi-rate robustness of isoconversional methods.

⚠️ **No reference E_a in the source sheet.** Treat as a
*self-consistency* fixture — different methods must agree with each
other across the same α window, but there is no independent truth.

## Data files

| File | Columns | Notes |
|------|---------|-------|
| `rate_<β>.tsv` × 6 | `alpha`, `T_K` | β ∈ {1, 2.5, 5, 10, 20, 40} K/min |

α grid: 0.005 → 0.995, step 0.01.

## Applicable tests

* **Multi-rate stress test** — agreement between KAS, OFW, Vyazovkin
  on a 6-rate dataset within an interior α window.
* **Method consistency** — Vyazovkin vs Vyazovkin-AIC at α ∈ [0.1, 0.9]
  should match to ~5 kJ/mol on this benign curve.
""",
)


def extract_epoxy() -> None:
    """Epoxy sheet has an irregular header layout; extract manually.

    Sheet layout (row 0 = header rates, row 1 = column labels, row 2+ = data):
        col 1: α
        col 2: T at 2.5 K/min
        col 3: T at 5 K/min
        col 4: T at 10 K/min
        col 6: Ea, kJ/mol (independent reference)
    """
    wb = load_workbook(VYAZOVKIN_XLSX, data_only=True)
    sheet = wb[_EPOXY.sheet]
    out = OUT / _EPOXY.folder
    rates_to_T_col = {2.5: 2, 5.0: 3, 10.0: 4}
    for beta, T_col in rates_to_T_col.items():
        a, T = _read_pairs(sheet, c_alpha=1, c_T=T_col)
        _save_tsv(out / f"rate_{beta}.tsv", ["alpha", "T_K"], [a, T])
    a_ref, ea_ref = _read_reference_Ea(sheet, alpha_col=1, ea_col=6)
    _save_tsv(out / "reference_Ea.tsv", ["alpha", "Ea_kJ_per_mol"], [a_ref, ea_ref])
    _write_readme(out / "README.md", _EPOXY.readme)
    wb.close()


def extract_skm() -> None:
    """SKM sheet: row 0 has labels (`a`, `T,K`), row 1 has rates 1–40, data row 2+."""
    wb = load_workbook(RASCHET_XLSM, data_only=True, keep_vba=True)
    sheet = wb[_SKM.sheet]
    out = OUT / _SKM.folder
    rates_to_T_col = {1.0: 2, 2.5: 3, 5.0: 4, 10.0: 5, 20.0: 6, 40.0: 7}
    for beta, T_col in rates_to_T_col.items():
        a, T = _read_pairs(sheet, c_alpha=1, c_T=T_col, start_row=3)
        _save_tsv(out / f"rate_{beta}.tsv", ["alpha", "T_K"], [a, T])
    _write_readme(out / "README.md", _SKM.readme)
    wb.close()


def extract_dapot_rate_1() -> None:
    """Recover β = 1 K/min for DAPOT from `3 скорости ДАПОТ` sheet.

    This rate is missing from the ДАПОТ sheet in Вязовкин.xlsx; users
    typically need the 4-rate dataset for jackknife uncertainty (n ≥ 3
    runs required after leave-one-out).
    """
    wb = load_workbook(RASCHET_XLSM, data_only=True, keep_vba=True)
    sheet = wb["3 скорости ДАПОТ"]
    # Sheet layout: col 1 = α, col 2 = T at 1 K/min (header rows 0..2).
    a, T = _read_pairs(sheet, c_alpha=1, c_T=2, start_row=4)
    out = OUT / "dapot"
    _save_tsv(out / "rate_1.0.tsv", ["alpha", "T_K"], [a, T])
    wb.close()


# ---------- xlsx cross-check extractors ----------

def extract_zplots() -> None:
    """`Сводная таблица.xlsx :: Z-plots` — precomputed Z(α) for ABPOT/DAPOT at 5 K/min.

    Sheet columns (1-indexed):
      ABPOT 5 K/min:  1=α, 2=T, 3=dα/dt, 4=Z_norm (= dα/dt·T² ÷ that quantity at α=0.5)
      DAPOT 5 K/min: 10=α, 11=T, 12=dα/dt, 13=Z_norm

    The xlsx "Z-plot" column 4 / 13 is *already normalized* so that
    Z(α=0.5) = 1 — confirmed by inspection: the value at α≈0.505 equals
    1.0 exactly. Use it directly.
    """
    wb = load_workbook(SVODKA_XLSX, data_only=True)
    sheet = wb["Z-plots"]

    def _extract(c_alpha, c_T, c_rate, c_z_norm, label):
        alpha, T, rate, z_norm = [], [], [], []
        for r in range(3, sheet.max_row + 1):
            a = sheet.cell(r, c_alpha).value
            t = sheet.cell(r, c_T).value
            d = sheet.cell(r, c_rate).value
            zz = sheet.cell(r, c_z_norm).value
            if a is None or t is None or d is None or zz is None:
                continue
            try:
                alpha.append(float(a))
                T.append(float(t))
                rate.append(float(d))
                z_norm.append(float(zz))
            except (TypeError, ValueError):
                continue
        out = OUT / "xlsx_crosscheck"
        _save_tsv(
            out / f"zplot_{label}_5kpm.tsv",
            ["alpha", "T_K", "dalpha_dt", "Z_norm_xlsx"],
            [np.array(alpha), np.array(T), np.array(rate), np.array(z_norm)],
        )

    _extract(1, 2, 3, 4, "abpot")
    _extract(10, 11, 12, 13, "dapot")
    wb.close()


def extract_A_calculated() -> None:
    """`Сводная таблица.xlsx :: A calculated` — precomputed A(α) for ABPOT at rates 1, 2.5, 5.

    Each rate-block has 6 columns:
        α, T_K, dα/dt, Evya (kJ/mol), A, (blank).
    Long-format output for easier consumption by tests.
    """
    wb = load_workbook(SVODKA_XLSX, data_only=True)
    sheet = wb["A calculated"]
    rate_starts = {1.0: 1, 2.5: 7, 5.0: 13}  # leftmost column of each rate-block
    rows: list[tuple[float, float, float, float, float, float]] = []
    for beta, c0 in rate_starts.items():
        for r in range(4, sheet.max_row + 1):
            a = sheet.cell(r, c0).value
            T = sheet.cell(r, c0 + 1).value
            d = sheet.cell(r, c0 + 2).value
            E = sheet.cell(r, c0 + 3).value
            A = sheet.cell(r, c0 + 4).value
            if any(v is None for v in (a, T, d, E, A)):
                continue
            try:
                rows.append((beta, float(a), float(T), float(d), float(E), float(A)))
            except (TypeError, ValueError):
                continue
    rows.sort(key=lambda row: (row[0], row[1]))
    arr = np.array(rows)
    out = OUT / "xlsx_crosscheck"
    _save_tsv(
        out / "A_calculated_abpot.tsv",
        ["rate_K_per_min", "alpha", "T_K", "dalpha_dt", "Evya_kJ_per_mol", "A_xlsx_per_sec"],
        [arr[:, i] for i in range(6)],
    )
    wb.close()


def extract_reaction_order() -> None:
    """`Сводная таблица.xlsx :: Проверка порядка` — precomputed
    y(n, α) = ln(dα/dT) − n · ln(1 − α) for n ∈ {1, 2, 3, 0.7}.

    ABPOT block (5 K/min): columns 1=α, 2=T, 3=dα/dT, 4=1/T, 5=y(n=1),
        6=y(n=2), 7=y(n=3), 8=y(n=0.7).
    DAPOT block (5 K/min): 11=α, 12=T, 13=1/T, 14=dα/dT, 15=y(1),
        16=y(2), 17=y(3), 18=y(0.7).
    Note the column-order difference: ABPOT has (dα/dT, 1/T) while
    DAPOT has (1/T, dα/dT).
    """
    wb = load_workbook(SVODKA_XLSX, data_only=True)
    sheet = wb["Проверка порядка"]

    def _read_block(c_alpha, c_T, c_dadT, c_invT, c_y1, c_y2, c_y3, c_y07, label):
        rows = []
        for r in range(7, sheet.max_row + 1):
            cells = [
                sheet.cell(r, c).value
                for c in (c_alpha, c_T, c_dadT, c_invT, c_y1, c_y2, c_y3, c_y07)
            ]
            if any(v is None for v in cells):
                continue
            try:
                rows.append([float(v) for v in cells])
            except (TypeError, ValueError):
                continue
        if not rows:
            return
        arr = np.array(rows)
        out = OUT / "xlsx_crosscheck"
        _save_tsv(
            out / f"reaction_order_{label}.tsv",
            ["alpha", "T_K", "dalpha_dT", "inv_T", "y_n1", "y_n2", "y_n3", "y_n0.7"],
            [arr[:, i] for i in range(8)],
        )

    # ABPOT: 5K/min block in columns 1-8.
    _read_block(1, 2, 3, 4, 5, 6, 7, 8, "abpot")
    # DAPOT: 5K/min block — alpha=11, T=12, invT=13, dα/dT=14, then y(n)=15..18.
    _read_block(11, 12, 14, 13, 15, 16, 17, 18, "dapot")
    wb.close()


def write_crosscheck_readme() -> None:
    _write_readme(
        OUT / "xlsx_crosscheck" / "README.md",
        """
# xlsx cross-check fixtures

Source: `theory/extra/Сводная таблица.xlsx` sheets `Z-plots`,
`A calculated`, `Проверка порядка`.

These tables contain values computed by the original Excel formulas
that motivated each algorithm. Used as **bit-exactness cross-checks**
against our re-implementations: our code reads (α, T, dα/dt) from the
same rows and must reproduce the precomputed Z / A / y(n) numbers.

## Data files

| File | Used by |
|------|---------|
| `zplot_abpot_5kpm.tsv`, `zplot_dapot_5kpm.tsv` | [`tests/test_xlsx_crosscheck_zplot.py`](../../../tests/test_xlsx_crosscheck_zplot.py) — Z(α) and normalized Z(α)/Z(0.5) at 5 K/min vs our `master_plot.py` |
| `A_calculated_abpot.tsv` | [`tests/test_xlsx_crosscheck_preexp.py`](../../../tests/test_xlsx_crosscheck_preexp.py) — A(α, run) under f(α) = 1 − α (F1) vs our `preexponential.py` |
| `reaction_order_abpot.tsv`, `reaction_order_dapot.tsv` | [`tests/test_xlsx_crosscheck_reaction_order.py`](../../../tests/test_xlsx_crosscheck_reaction_order.py) — y(n; α, T) = ln(dα/dT) − n · ln(1 − α) at n ∈ {1, 2, 3, 0.7} vs our `reaction_order.py` |

The cross-check tests use *the xlsx's own (α, T, dα/dt)* as input so
that any disagreement is attributable to a difference in computation,
not in extraction.
""",
    )


def write_root_readme() -> None:
    _write_readme(
        OUT / "README.md",
        """
# Reference workbooks

Source: independent Excel workbooks (`theory/extra/*.xlsx`, gitignored).
This directory contains the *extracted* numerical data — no .xlsx kept
in version control.

Material data is split per folder; cross-check tables that originate
from xlsx-internal computations live in `xlsx_crosscheck/`.

| Folder | Material | Heating rates (K/min) | Has reference E_a? |
|--------|----------|-----------------------|--------------------|
| [`abpot/`](abpot/) | ABPOT polymer | 0.2, 1.0, 2.5, 5.0 | ✓ |
| [`dapot/`](dapot/) | DAPOT polymer | 1.0, 2.5, 5.0, 10.0 | ✓ |
| [`epoxy_pv15/`](epoxy_pv15/) | Epoxy PV15-0,5 | 2.5, 5.0, 10.0 | ✓ |
| [`skm_prepreg/`](skm_prepreg/) | SKM preimpregnated | 1.0, 2.5, 5.0, 10.0, 20.0, 40.0 | ✗ |
| [`xlsx_crosscheck/`](xlsx_crosscheck/) | precomputed Z, A, y(n) | n/a | n/a |

## Re-extraction

If the source xlsx files are restored under `theory/extra/`, regenerate
everything with::

    .venv/bin/python scripts/extract_reference_fixtures.py

The script is idempotent — re-running overwrites existing TSV files
with identical content (bit-exact under the same xlsx).
""",
    )


def main() -> None:
    if not SRC.exists():
        raise SystemExit(
            f"Source folder {SRC} not present — restore xlsx files before running."
        )
    OUT.mkdir(parents=True, exist_ok=True)
    extract_material(VYAZOVKIN_XLSX, _ABPOT)
    extract_material(VYAZOVKIN_XLSX, _DAPOT)
    extract_dapot_rate_1()
    extract_epoxy()
    extract_skm()
    extract_zplots()
    extract_A_calculated()
    extract_reaction_order()
    write_crosscheck_readme()
    write_root_readme()
    print(f"All fixtures written to {OUT}")


if __name__ == "__main__":
    main()
